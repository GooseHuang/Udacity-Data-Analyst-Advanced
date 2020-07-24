# -*- coding: utf-8 -*-
"""
数据框处理工具包：
本程序包是日常常用函数的集合，使用时将以下代码加入到程序中即可使用程序包

部分程序会有多个版本，根据需要尝试和替换

import sys
sys.path.insert(0, 'D:\python_toolkit')   #将数据包所在路径临时加入系统路径
import toolkit as tk

import pandasql as pdsql
pysql = lambda q:pdsql.sqldf(q,globals())
import os
import pandas as pd

"""


#安装MySQLdb时使用
# pip install mysql-connector-python-rf

import os
import datetime
import pandas as pd
#import mysql.connector as sql
import re
#import MySQLdb


#多列合一函数，保留所选择列的不为空的记录
def trans_col(df_todo,col_to_append,new_col):
    """
    将宽数据转换为长数据
    
    注：原始合并的列将不存在，会生成为新的一列，原列中空值会被删除
    
    输入为：
    df_todo 需要合并的dataframe
    col_to_append  dataframe中需要合并、纵向追加的列
    new_col  生成的新列名
    
    old df:
    
      日期       品种1      品种2       品种3
    2017-11-2    苹果       香蕉        梨
    2017-11-3    橘子       葡萄        哈密瓜
    2017-11-4    苦瓜        -           -  
    
    new df:
    
      日期      水果品种
    2017-11-2    苹果
    2017-11-2    香蕉
    2017-11-2    梨
    2017-11-3    橘子 
    2017-11-3    葡萄 
    2017-11-3    哈密瓜
    2017-11-4    苦瓜 
    """
    
    data = df_todo.copy()
    #获取要保留的列（col_to_append之外所有的列）
    col_to_keep = list(set(data.columns).difference(set(col_to_append)))
    
    
    df_wanted = pd.DataFrame(columns=col_to_keep+[new_col])
    
    for col in col_to_append:
        df_wanted.rename(columns={df_wanted.columns[-1]: col},inplace=True)  
        
        df_wanted = df_wanted.append(data.ix[data[col]!="",col_to_keep+[col]])
    df_wanted.columns = col_to_keep+[new_col] 

	#重新排列列顺序
    col_list = []
    for col in data.columns:
        if col in df_wanted.columns:
            col_list.append(col)
    col_list.append(new_col)
    return df_wanted.loc[:,col_list]




#分列函数 按分隔符分列，新增的列会以num_1,num_2...命名，置于该列之后
#将trans设置为True，可以纵向拼接  
def seperate_col(df_todo,col,sep="，",trans=False): 
    
    """
    将某列按分隔符分开，生成多列，自动命名为num_1,num_2...，
    
    注：原始列会被予以保留
    
    输入为：
    df_todo 需要处理的dataframe
    col     dataframe中需要分隔的列，操作对象需要转化成str
    sep     分隔符，默认分隔符为中文逗号'，'
    trans   是否需要将新生成的列纵向拼接，形成一列长数据，默认为False
    
    old df:
    
      日期             品种
    2017-11-2    苹果，香蕉，梨
    2017-11-3    橘子，葡萄，哈密瓜
    2017-11-4    苦瓜  
          
    new df:
        
    trans = False
    
      日期               品种               num_1       num_2       num_2
    2017-11-2  苹果，香蕉，梨  苹果          苹果        香蕉        梨
    2017-11-3  橘子，葡萄，哈密瓜  橘子      橘子        葡萄        哈密瓜
    2017-11-4  苦瓜                         苦瓜        -           -  
    
    trans = True    
    
      日期      品种
    2017-11-2    苹果
    2017-11-2    香蕉
    2017-11-2    梨
    2017-11-3    橘子 
    2017-11-3    葡萄 
    2017-11-3    哈密瓜
    2017-11-4    苦瓜 
    
    """  
    
    data = df_todo.copy()
    
    #转换为list
    data[col] = data[col].str.split(sep)
    num_col = 'num'
    data[num_col] = data[col].apply(len)
    
    for i in range(max(data[num_col])):
        name=num_col + '_'+str(i+1)
        data[name]=data[col].apply(lambda x:x[i] if (len(x)>i) else '')
    
    num_list=[num_col+'_'+str(i+1) for i in range(max(data[num_col]))]
    data.drop([col,num_col],axis=1,inplace=True)
    if trans:
        data = trans_col(data,num_list,col)
    return data

#合并列
def concat_columns(df_todo,col_list,new_col,sep='，',drop_col=False):
        
    """
    将多列按分隔符合并，生成新列，并可以选择是否保留原有的列
    
    注：原始列会被予以保留
    
    输入为：
    df_todo      需要处理的dataframe
    col_list     dataframe中需要合并的列
    new_col      合并成新列的列名
    sep          分隔符，默认分隔符为中文逗号'，'
    drop_col     是否删除原有的列，默认为False
    
    old df:
        
      日期        水果1       水果2       水果3
    2017-11-2     苹果        香蕉        梨
    2017-11-3     橘子        葡萄        哈密瓜
    2017-11-4     苦瓜        -           -  
    
    col_list = ['水果1','水果2','水果3']
    new_col = '品种'

    new df:
        
    drop_col = False
    
      日期         水果1       水果2       水果3         品种
    2017-11-2      苹果        香蕉        梨         苹果，香蕉，梨  苹果 
    2017-11-3      橘子        葡萄        哈密瓜      橘子，葡萄，哈密瓜  橘子
    2017-11-4      苦瓜        -           -          苦瓜               
    
    drop_col = True    
    
    
      日期             品种
    2017-11-2    苹果，香蕉，梨
    2017-11-3    橘子，葡萄，哈密瓜
    2017-11-4    苦瓜  
    
    """  
    
    ###########
    #df_todo = df_xidesheng
    #col_list = ['LS锁维修类型','2.0S锁维修类型','HS锁维修类型','FS锁维修类型']
    #sep ='，'
    #new_col='锁维修类型'
    ###########
    
    data = df_todo.copy()
    data[new_col]=""
    for col in col_list:
        data[new_col]= data[new_col]+data[col].apply(lambda x:sep if x != '' else '') +data[col]
    
    data[new_col] = data[new_col].apply(lambda x:re.sub("^"+sep,'',x))
    if drop_col:
        for col in col_list:
            del data[col]
    return data
 



def multicol_append_to_one(df_todo,col_to_keep,col_to_add,new_col):
    """
    将df中的多列纵向追加合并成一列，宽数据变成长数据，可以选择要保留的列
    
    输入：
    df_todo                 需要合并的df
    col_to_keep             df中其他需要保留的列
    col_to_add              df中需要纵向合并的列
    new_col                 合并成新列的命名
    
    
    样例
    old df:
      日期               果园               水果1       水果2       水果3
    2017-11-2            蜀国               苹果        香蕉        梨
    2017-11-3            魏国               橘子        葡萄        哈密瓜
    2017-11-4            吴国               苦瓜        -           -  
        
    col_to_keep = ['日期','果园']
    col_to_add  = ['水果1','水果2','水果3']
    new_col = '水果'
    
    new df:
      日期               果园               水果
    2017-11-2            蜀国               苹果
    2017-11-3            魏国               橘子
    2017-11-4            吴国               苦瓜
    2017-11-2            蜀国               香蕉
    2017-11-3            魏国               葡萄 
    2017-11-2            蜀国               梨
    2017-11-3            魏国               哈密瓜
        
    """
    df = df_todo.copy()
    
    df_wanted = pd.DataFrame(columns=col_to_keep+[new_col])
    
    for col in col_to_add:
        df_temp = df.loc[:,col_to_keep+[col]]
        df_temp.rename(columns={col:new_col},inplace=True)
        df_wanted = df_wanted.append(df_temp)
    
    #df_wanted = df_wanted.loc[:,col_to_keep+[new_col]]
    
    df_wanted = subDataframe(df_wanted,new_col,"",operator="!=")
    
    return df_wanted	



"""
#按多列取出distinct行的函数
def get_distinct_combination(df_todo,col_to_mark):
    
    ###########
#    df_todo = df_all.copy()
#    col_to_mark = [Biketype,Failure]
    
    ###########
    
    df = df_todo.copy()
    select = " SELECT "
    comma = ","
    froma = " FROM "
    groupby = " GROUP BY "
    table = 'df'
    
    import pandasql as pdsql
    pysql = lambda q:pdsql.sqldf(q,locals())
    
    query = select + comma.join(col_to_mark) + froma + table + groupby + comma.join(col_to_mark)
    
    df_wanted = pysql(query)
    
    return df_wanted
"""



#日期转换，供SQL语句使用
def date_quotes(date):
    #date='2017-12-23'
    return "'"+date+"'"
	

#dataframe因子替换    
def replaceFactor(df_todo,col_name,factor_dict = {}):
    
    """
    将dataframe目标列中的值，按照输入字典的逻辑替换
    
    输入为：
    df_todo       需要处理的dataframe
    col_name      将要采取替换操作的列
    factor_dict   包含替换逻辑的词典 '旧元素':'新元素'
    
    old df:
    
      日期             品种
    2017-11-2         烂苹果
    2017-11-3         烂葡萄
    2017-11-4          苦瓜
    
    factor_dict = {
                    '烂苹果':'好苹果',
                    '烂葡萄':'好葡萄',
                    '苦瓜':'香瓜'
                    }
          
    new df:
        
      日期             品种
    2017-11-2         好苹果
    2017-11-3         好葡萄
    2017-11-4          香瓜        

    
    """   
    
    df = df_todo.copy()
    for key,value in factor_dict.items():
        df.ix[df[col_name]==key,col_name] = value
    return df

#输入list，返回便于加工成字典的格式
def return_dict_format(name_list):
    
    """
    将list加工成便于输入的字典格式，返回str
    
    输入为：
    name_list     需要处理成字典的list

    
    old list:
    
        ['好苹果','烂苹果','苦瓜']
          
    new string:
        
        "{'好苹果': '', '烂苹果': '', '苦瓜': ''}"

    
    """   
    input_dict = {}
    for key in name_list:
        input_dict[key] = ''
    
    return str(input_dict)	
	
	
#读取csv，以str格式
def read_csv_in_str(path,encoding='gb2312',sep="\t",header=0):
    """
    将csv按照str的格式读入
    
    输入为：
    path          文件路径或文件名
    encoding      csv文件的编码格式
    sep           csv文件的分隔符
    header        列名所在的行数，默认为0
    
    """
    column_list = []
    df_column = pd.read_csv(path,encoding = encoding,sep=sep,header=header).columns
    for i in df_column:
        column_list.append(i)
    data = pd.read_csv(path,converters={col: str for col in column_list},encoding = encoding,sep=sep,header=header)
    return data

#读取Excel，以str格式,可指定fill_na为False，以便不填充空值
def read_excel_in_str(path,sheetname='Sheet1',fill_na=""):
    """
    将excel按照str的格式读入
    
    输入为：
    path          文件路径或文件名
    sheetname     为worksheet的名字
    fill_na       是否替换空值，默认替换为""，若不处理空值，可以设置为False
    
    """    
    
    
    column_list = []
    df_column = pd.read_excel(path,sheetname ).columns
    for i in df_column:
        column_list.append(i)
    data = pd.read_excel(path,sheetname,converters={col: str for col in column_list})
    if fill_na != False:
        data.fillna(fill_na,inplace=True)
    return data


#读取Excel，默认读取第一张表
def read_excel_in_strV2(path,sheetname='',fill_na=""):
    """
    将excel按照str的格式读入，默认读取第一张表
    注：openyxl函数库限制，无法读取xls
    
    输入为：
    path          文件路径或文件名
    sheetname     为worksheet的名字，如果未指定，则默认读取第一张表
    fill_na       是否替换空值，默认替换为""，若不处理空值，可以设置为False
    
    """    
    from openpyxl import load_workbook

    wb = load_workbook(path)
    #获取sheetname
    sheets = wb.get_sheet_names()
    sheet0 = sheets[0]
     
    del wb
    
    if sheetname == '':
        sheetname = sheet0
        
    
    
    column_list = []
    df_column = pd.read_excel(path,sheetname ).columns
    for i in df_column:
        column_list.append(i)
    data = pd.read_excel(path,sheetname,converters={col: str for col in column_list})
    if fill_na != False:
        data.fillna(fill_na,inplace=True)
    return data



#读取Excel，默认读取第一张表,自动识别sheet名，兼容xls
def read_excel_in_strV3(path,sheetname='',header=0,fill_na="",skiprows=0):
    
    """
    将excel按照str的格式读入，默认读取第一张表，兼容xls
    
    输入为：
    path          文件路径或文件名
    sheetname     为worksheet的名字，如果未指定，则默认读取第一张表
    header        标题所在的位置
    fill_na       是否替换空值，默认替换为""，若不处理空值，可以设置为False
    
    """    
    
    import xlrd
    xls = xlrd.open_workbook(path, on_demand=True)
    
    sheets = xls.sheet_names()
    sheet0 = sheets[0]
     
    del xls
    
    if sheetname == '':
        sheetname = sheet0
        
    
    
    column_list = []
    df_column = pd.read_excel(path,sheetname=sheetname,header=header,skiprows=skiprows).columns
    for i in df_column:
        column_list.append(i)
    data = pd.read_excel(path,sheetname=sheetname,header=header,skiprows=skiprows,converters={col: str for col in column_list})
    if fill_na != False:
        data.fillna(fill_na,inplace=True)
    return data


#读取excel表头，并去除脏数据
def read_table_name(table):
    """
    有些excel表头中存在脏的字符，如'\ufeff','\t',' '等等，该程序可以清洗表头，返回清洗好的表头
    
    输入为：
    table         文件路径或文件名
    """
    
    import xlrd
    xls = xlrd.open_workbook(table, on_demand=True)
    
    #sheets = xls.sheet_names()
    #sheet0 = sheets[0]
         
    sheet = xls.sheet_by_index(0)
    
    columns = [sheet.cell_value(0, col) for col in range(sheet.ncols)]
    
    for ugly_word in ["\ufeff"," ",""]:
        columns = [str(x).replace(ugly_word,"") for x in columns]
    del xls
    return columns



#读取Excel，读取所有的表，可指定标题位置
def read_all_sheets_in_str(path,header=0,fill_na=""):
    
    """
    将excel按照str的格式读入，自动拼接所有表格成为一个dataframe，各worksheet表头需一致
    
    输入为：
    path          文件路径或文件名
    header        标题所在的位置
    fill_na       是否替换空值，默认替换为""，若不处理空值，可以设置为False
    """  
    
    xl = pd.ExcelFile(path)
    #获取sheetname
    sheets = xl.sheet_names

    column_list = []
    df_column = pd.read_excel(path,sheetname=sheets[0],header=header).columns                  
    for i in df_column:
        column_list.append(i)
     
    df_all = pd.DataFrame(columns=df_column)
    for sheet in sheets: 
        print(sheet)
        df_temp = pd.read_excel(path,sheetname=sheet,header=header,converters={col: str for col in column_list})
        df_all = df_all.append(df_temp)
    if fill_na != False:
        df_all.fillna(fill_na,inplace=True)
    return df_all


#去除string中不需要的字符
def cleanText(x,word_list):
    """
    去除string中不需要的字符
    
    输入为：
    x             目标string
    word_list     需要去除的字符
    
    input:
        
        x = "\ufeff apple  \t"
        word_list = ['\ufeff','\t',' ']
    
    output:
    
        "apple"        
    """  
    try:
        x = str(x)
    except Exception as err:
        print(x,err)
        return ""    
    for word in word_list:
        x = x.replace(word,"")            
    return x 


def replaceText(string,rule):
    """
    去除string中不需要的字符
    
    输入为：
    x             目标string
    rule          字符替换规则，可以是dict、list、string
    
    input:
        
        x = "\ufeff apple  \t"
        word_list = ['\ufeff','\t',' ']
    
    output:
    
        "apple"        
    """  
    if type(rule)==list:
        for item in rule:
            string = string.replace(item,"")
    elif type(rule)==dict:
        for key,value in rule.items():
            string = string.replace(key,value)
    else:
        string = string.replace(rule,"")
    return string




#设置数据库基本信息，配合conn = set_connecter()使用
def set_connecter(db,host='localhost', user='root',passwd='mobike123456',charset='utf8'):
    """
    设置数据库基本信息，返回conn供进一步数据库操作使用
    输入：
    db                 需要操作的数据库名
    host               主机地址，默认'localhost'
    user               用户名，默认'root'
    passwd             密码，默认'mobike123456'
    charset            读写字符集，默认'utf8'
    """
    conn = MySQLdb.connect(host=host, user=user,passwd=passwd,db=db,charset=charset)
    return conn
	
#插入df到数据库，df的列名和数据库中有的需要一一对应
def insert_df_into_database(df_todo,table,conn,path=os.getcwd()):
    """
    插入df到数据库，df中的列名在数据库中必须全部存在，自动按名称匹配
    输入：
    df_todo            需要上传的dataframe
    table              想要插入的表格
    conn               数据库操作对象
    path               日志生成路径，如果出错会在该路径下生成错误报告，默认程序当前路径
    """
    
    data = df_todo.copy()
    cursor = conn.cursor()
    error_log = []
    for i in range(data.shape[0]):
        ######
        #i = 1
        #table = 'diaodubiao_test'
        ######
        query = "INSERT INTO " +table + str.replace(str(tuple(data.columns)),"'","") + " VALUES " 
                                            
        values = str.replace(str(tuple(data.iloc[i])),"nan","''")

        query = query + values
        
        
        try:
            cursor.execute(query)
            print(i,'Done!')
        except Exception as err:
            error_log.append([i,err])
            print(i,err)
        #print(query)
    conn.commit()
    
    
    
    if len(error_log)>0:
        outdir = path+"\\"+table+"_upload_error.txt"
        with open(outdir, "w") as output:
            output.write("error_number: "+str(len(error_log))+"\n")
            for value in error_log:
                output.write(str(value)+"\n")
            output.write("error_number: "+str(len(error_log)))
        print("good number: " + str(data.shape[0]-len(error_log)))
        print("error number: "+str(len(error_log)))
        print("error dir: "+outdir)
        
        #写入日志文件夹
        outdir = "D:\\error_log\\"+table+"_upload_error.txt"
        with open(outdir, "w") as output:
            output.write("error_number: "+str(len(error_log))+"\n")
            for value in error_log:
                output.write(str(value)+"\n")
            output.write("error_number: "+str(len(error_log)))
        
        
    else:
        print("\nExcuted without error.\n")


#数据库上传函数，按最大序号上传：
def get_insert_df_by_max_num(df_todo,target_table,num_col,database,host='localhost',user='root', password='mobike123456'):
    
    """
    从df中截取需要插入数据库中的部分，数据库和df需要有一列名称相同的序号列。
    插入之后可以保证数据的序号是连贯的，并且避免重复插入
    
    输入：
    df_todo            需要插入的dataframe        
    target_table       想要插入的表格
    num_col            df_todo中序号列所在的行    
    database           需要操作的数据库名
    host               主机地址，默认'localhost'
    user               用户名，默认'root'
    passwd             密码，默认'mobike123456'
    """
    
    ##############
#    df_todo = data
#    target_table = 'ganyu_table'
#    num_col = 'xuhao'
#    database = 'mobike_ganyu'
#    host='localhost'
#    user='root'
#    password='mobike123456'
    ############
    #目标df
    df=df_todo.copy()
    #目标数据库
    db_connection = sql.connect(host=host,database=database, user=user, password=password)
    #获取表单的最大序号
    query="SELECT MAX(CAST( "+num_col+" AS SIGNED)) AS title_no FROM "+target_table+";"
                     
    max_num = pd.read_sql(query,con=db_connection).iloc[0,0]
    
#    "select max(col) from (select to_number("+num_col+") as col from "+target_table+") b"
     
    #查看df中，是否有大于序号的值，获取到的为df_wanted
    if str(max_num+1) in list(df[num_col]):
        print("Great! We have what we wanted. We didn't miss any value!")
    else:
        print("Sorry! We didn't find what we wanted. We are missing value!")
        print("Max value in database: ",max_num)
        print("Value Range in df: ",df[num_col].min(),'-',df[num_col].max())
        print("We don't have anything to insert!")
        return None
        
    df_wanted = df.ix[df[num_col]>=str(max_num+1),:] 
    return df_wanted
	

#读取数据库，生成dataframe
def pandas_mysql(query,database,host='localhost', user='root', password='mobike123456'):
    """
    直接利用read_sql函数，对数据库进行查询，生成dataframe
    """
    db_connection = sql.connect(host=host, database=database, user=user, password=password)
    df = pd.read_sql(query, con=db_connection)
    return df

#获取数据库列名
def get_table_names(table_name,database):
    """
    获取数据库中表格的全部列名，返回一个list
    """
    return list(pandas_mysql("SELECT * FROM "+table_name+" LIMIT 0 ",database).columns)


#获取excel sheet名
def get_sheetnames(path):
    """
    获取excel sheet名
    """
    xl = pd.ExcelFile(path)
    sheets = xl.sheet_names
    return sheets



	
	
#去除元素中不想要的字符串
def replace_unwanted_text_in_col(df_todo,col,text_dict =dict()):
    """
    去除df某列元素中不想要的字符，替换成特定字符
    注：字符只要存在于目标字符串中，全部会被替换掉
    
    输入：
    df_todo       需要处理的df
    col           需要处理的列
    text_dict     替换规则字典
    
    old df:
    
      日期             品种
    2017-11-2         \t烂苹果
    2017-11-3         @烂葡萄
    2017-11-4          *苦瓜
    
    text_dict = {
                    '\t':'',
                    '@':'',
                    '*':'好多'
                    }
          
    new df:
        
      日期             品种
    2017-11-2         烂苹果
    2017-11-3         烂葡萄
    2017-11-4         好多苦瓜        
    
    """
    data = df_todo.copy()
    for key,value in text_dict.items():
           data[col] = data[col].apply(lambda x:x.replace(key,value))
    return data


	
#分级	
#保留当天维修最高分级的车辆   因为同一天会修很多车，加上按照时间和仓库的排序方式，只需要关注该车辆ID即可
def select_high(df_todo,Date,Warehouse,ID,Devision):
    """
    主要用于维修系统，在车辆定级之后，筛选出当天，同一仓库，同一个id车辆的最高分级记录（只保留一个）
    注：适用于早期mola的维修分级
    
    输入：
    df_todo       需要处理的df
    Date          日期列名
    Warehouse     仓库列名    
    ID            车辆ID列名
    Devision      故障分级，需要以数字格式输入
       
    old df：
    日期          仓库         ID           分级
    2017-11-2    横岗仓库    7558495128       0
    2017-11-2    横岗仓库    7558495128       3 
    2017-11-3    横岗仓库    7558495128       2
    2017-11-2    喜德盛仓库  7895621545       1
    
    new df：
    日期          仓库         ID           分级
    2017-11-2    横岗仓库    7558495128       3 
    2017-11-3    横岗仓库    7558495128       2
    2017-11-2    喜德盛仓库  7895621545       1    
    
    """
    
    data = df_todo.copy()
    data= data.sort_values([Date,Warehouse,ID,Devision],ascending=[True,True,True,False])
    data.index = range(data.shape[0])
    high_list = []
    i = 0
    while i < data.shape[0]:
        high_list.append(i)
        try :
            while data[ID][i+1] == data[ID][i]:
                i += 1
        except Exception as err:
            print(i,err)
        i += 1
        print(i)
    
    other_list = list(set(data.index).difference(set(high_list)))  
    data_high = data.ix[high_list,:]
    data_others=data.ix[other_list,:]
    return data_high,data_others

	
def seperate_keep_distinct(df,ID = "维修车辆编号",Warehouse = '仓库',Date = '修车日期',n=7,keep_early_date=True):
    
    """
    主要用于维修系统，将数据按照同一仓库、车辆编号、n天内的规则去重，返回去重后的唯一值，可以选择保持较早的或较晚的值，
    返回两个dataframe，一个是去重后的，一个是重复值
    注：适用于早期mola的维修分级
    
    输入：
    df_todo           需要处理的df
    Date              日期列名
    Warehouse         仓库列名    
    ID                车辆ID列名
    n                 去除n天内重复出现的值
    keep_early_date   是否保留较早出现的值，默认为True
       
    old df：
    日期          仓库         ID          
    2017-11-2    横岗仓库    7558495128    
    2017-11-2    横岗仓库    7558495128     
    2017-11-3    横岗仓库    7558495128     
    2017-11-2    喜德盛仓库  7895621545      
    
    keep_early_date=True
    
    new df：
    
    df_good
    
    日期          仓库         ID           
    2017-11-2    横岗仓库    7558495128         
    2017-11-2    喜德盛仓库  7895621545  

    df_duplicate  
    
    日期          仓库         ID          
    2017-11-2    横岗仓库    7558495128        
    2017-11-3    横岗仓库    7558495128      
    
    """    
    
    ###########################
    #测试
    #ID = "维修车辆编号"
    #Warehouse = '仓库'
    #Date = '修车日期'
    #n=7
    #############################
    #日期转换成时间格式
    data = df.copy()
    data[Date] = data[Date].str.extract('(\d\d\d\d-\d\d-\d\d)', expand=False)
    data[Date] = pd.to_datetime(data[Date],format='%Y-%m-%d')
    data[Date] = data[Date].dt.to_period('D')

    #排序 
    data =  data.sort_values([ID,Warehouse,Date],ascending=[True,True,False])
    data.index= range(data.shape[0])


    #print("Begin!")
    #筛选,取出重复嫌疑值，保留唯一值
    duplicate_list = []
    if keep_early_date:  
        for i in range(0,data.shape[0]-1):
            #最后一个值肯定被保留（不可能被选到duplicate_list）
            if (data[ID][i+1] == data[ID][i] and
                data[Warehouse][i+1] == data[Warehouse][i] and
                data.ix[i,Date] - data.ix[i+1,Date] < n) :
                duplicate_list.append(i)
                #print(i) 
    else:
        for i in range(0,data.shape[0]-1):
            #第一个值肯定被保留（不可能被选到duplicate_list）
            if (data[ID][i+1] == data[ID][i] and
                data[Warehouse][i+1] == data[Warehouse][i] and
                data.ix[i,Date] - data.ix[i+1,Date] < n) :
                duplicate_list.append(i+1)
                #print(i+1)    
    
    #日期转换回str
    data[Date] = data[Date].astype(str)
    
    #好的数据的集合
    good_list = list(set(data.index).difference(set(duplicate_list)))  
 
    ########################
    #删除的重复值     
    df_duplicate = data.ix[duplicate_list,:]
    #######################
    #去重后的集合
    df_good = data.ix[good_list,:]
    ####################
    #print("Done!")
    return df_good,df_duplicate
	
	
	
def seprate_remove_all(df,ID = "维修车辆编号",Warehouse = '仓库',Date = '修车日期',n=7):
    """
    主要用于维修系统，将数据按照同一仓库、车辆编号、n天内的规则去重，删除全部的重复值
    返回两个dataframe，一个是去重后的，一个是重复值（任何在规则中出现两次以上的值都会被删除）
    注：适用于早期mola的维修分级
    
    输入：
    df                需要处理的df
    Date              日期列名
    Warehouse         仓库列名    
    ID                车辆ID列名
    n                 去除n天内重复出现的值
       
    old df：
    日期          仓库         ID          
    2017-11-2    横岗仓库    7558495128    
    2017-11-2    横岗仓库    7558495128     
    2017-11-3    横岗仓库    7558495128     
    2017-11-2    喜德盛仓库  7895621545      
    
    new df：
    
    df_good
    
    日期          仓库         ID                   
    2017-11-2    喜德盛仓库  7895621545  

    df_duplicate  
    
    日期          仓库         ID
    2017-11-2    横岗仓库    7558495128           
    2017-11-2    横岗仓库    7558495128        
    2017-11-3    横岗仓库    7558495128      
    
    """    
    ###########################
    #测试
    #ID = "维修车辆编号"
    #Warehouse = '仓库'
    #Date = '修车日期'
    #n=6
    #############################
    #日期转换成时间格式
    data = df.copy()
    data[Date] = data[Date].str.extract('(\d\d\d\d-\d\d-\d\d)', expand=False)
    data[Date] = pd.to_datetime(data[Date],format='%Y-%m-%d')
    data[Date] = data[Date].dt.to_period('D')

    #排序 
    data =  data.sort_values([ID,Warehouse,Date],ascending=[True,True,False])
    #缺失是从数字最大的日期排到最近的
    #data.sort_values([Date],ascending=[False])['修车日期']
    data.index= range(data.shape[0])





    print("Begin!")
    #筛选,取出所有重复嫌疑值，不保留
    duplicate_list = []
    for i in range(1,data.shape[0]-1):
        if (data[ID][i+1] == data[ID][i] and
            data[Warehouse][i+1] == data[Warehouse][i] and
            data.ix[i,Date] - data.ix[i+1,Date] < n) :
            duplicate_list.append(i)
            print(i)
        elif(data[ID][i-1] == data[ID][i] and
            data[Warehouse][i-1] == data[Warehouse][i] and
            data.ix[i,Date] - data.ix[i-1,Date] < n):
            duplicate_list.append(i)
            print(i)
    i = data.shape[0]-1
    if(data[ID][i-1] == data[ID][i] and
            data[Warehouse][i-1] == data[Warehouse][i] and
            data.ix[i,Date] - data.ix[i-1,Date] < n):
        duplicate_list.append(i)
        print(i)
    
    
    
    #日期转换回str
    data[Date] = data[Date].astype(str)
    
    #好的数据的集合
    good_list = list(set(data.index).difference(set(duplicate_list)))  
 
    ########################
    #删除的重复值     
    df_duplicate = data.ix[duplicate_list,:]
    #######################
    #去重后的集合
    df_good = data.ix[good_list,:]
    ####################
    print("Done!")
    return df_good,df_duplicate
	
	

    
#dataframe取子集	
def subDataframe(df_todo,col,value="",operator='=='):
    """
    依据dataframe某列的值取子集，一次只能筛选某列的某个值
    注：设置该函数是因为pandas内置方法不直观
    输入：
    df_todo      目标dataframe    
    col          dataframe中需要筛选特定值的列
    value        需要筛选出的值
    operator     运算符，可输入'==','!=','>','<'等
    
    输入：
    old df
        日期          仓库         ID          
    2017-11-2    横岗仓库    7558495128    
    2017-11-2    横岗仓库    7558495128     
    2017-11-3    横岗仓库    7558495128     
    2017-11-2    喜德盛仓库  7895621545   
    
    col = 'ID'
    value = '7558495128'
    operator = '=='
    
    new df
        日期          仓库         ID          
    2017-11-2    横岗仓库    7558495128    
    2017-11-2    横岗仓库    7558495128     
    2017-11-3    横岗仓库    7558495128     
    """
    
    df = df_todo.copy()
    if operator == '==':
        return df.ix[df[col]==value,:]
    elif operator == '!=' or operator == '<>':
        return df.ix[df[col]!=value,:]
    elif operator == '>=':
        return df.ix[df[col]>=value,:]
    elif operator == '>':
        return df.ix[df[col]>value,:]
    elif operator == '<=':
        return df.ix[df[col]<=value,:]
    elif operator == '<':
        return df.ix[df[col]<value,:]
    
    
#加强版，可取一列中在list中的值,只能判断是否相等
def subDataframeD1(df_todo,col,value_list,operator='=='):
    """
    subDataframe加强版，依据dataframe某列的值取子集，一次可以筛选某列的多个值
    
    输入：
    df_todo      目标dataframe    
    col          dataframe中需要筛选特定值的列
    value_list   需要筛选出的值的list
    
    样例：
    old df
        日期          仓库         ID          
    2017-11-2    横岗仓库    7558495128    
    2017-11-2    横岗仓库    7558495128     
    2017-11-3    横岗仓库    7558495128     
    2017-11-2    喜德盛仓库  7895621876
    2017-11-2    喜德盛仓库  7894875545  
    
    col = 'ID'
    value_list = ['7895621876','7894875545']
    
    new df
        日期          仓库         ID          
    2017-11-2    喜德盛仓库  7895621876
    2017-11-2    喜德盛仓库  7894875545  
    """
    
    if type(value_list) != list:
        value_list = []
        value_list.append(value_list)
    
    if operator=='==':
        data = pd.DataFrame(columns = df_todo.columns) 
        for value in value_list:
            df_temp = subDataframe(df_todo,col,value,operator)
            data = data.append(df_temp)
    elif operator=='!=' or operator=='<>' :
        data = df_todo.copy()
        for value in value_list:
            data = subDataframe(data,col,value,operator)
    return data	



#加强版，可取多列
def subDataframeV2(df_todo,value_dict,operator='=='):
    
    """
    subDataframe加强版，依据dataframe某列的值取子集，可以根据规则筛选多列的值
    
    输入：
    df_todo      目标dataframe    
    col          dataframe中需要筛选特定值的列
    value_dict   删选规则字典
    
    样例：
    old df
        日期          仓库         ID          
    2017-11-2    横岗仓库    7558495128    
    2017-11-2    横岗仓库    7558495128     
    2017-11-3    横岗仓库    7558495128     
    2017-11-2    喜德盛仓库  7895621876
    2017-11-2    喜德盛仓库  7894875545  
    
    col = 'ID'
    value_list = ['日期':'2017-11-3','ID':'7558495128']
    operator='=='
    
    new df
        日期          仓库         ID          
    2017-11-3    横岗仓库    7558495128 
    """
    ##########
#    df_todo =  df_ganyu_sum
#    value_dict = {'mark':key,'区域':the_quyu}  
    ########## 
    
    data = df_todo.copy()
    for col,value in value_dict.items():
        #############
#        col= 'mark'
#        value='故障车干预总数-运营'
        #############
        
        data = subDataframe(data,col,value,operator)
    return data		
		
#维修分级     横向分级		
def max_devision(df_todo,df_set,classify_by_A_others,col_to_classify,classify_by_B,classify_target,new_col,sep='，',keep_col=True):
    """
    本程序主要应对维修系统，将维修记录按相应规则分级，并保留最高分级。
    注：1.原始数据中一辆车往往包含多个故障（该程序也可以应对单一故障的情况），处理时会自动按分隔符分开；
        2.该程序可以拓展到其他多重依据分类的情况。
        3.对应新版的mola维修数据；
        4.分级标准里没有的都被划分为了零级
        5.程序会自动去除目标列（col_to_classify）中存在的空格
        6.df_todo原有的列会得到保留
        
    输入：
    df_todo                需要分级的dataframe，设为表A
    df_set                 包含分级规则的dataframe，设为表B
    classify_by_A_others   表A中除了要按分隔符分开的列（故障所在列），其他依据列的列名集
    col_to_classify        表A中需要按分隔符分开的列（故障所在列）
    classify_by_B          表B中包含分级依据的列名集
    classify_target        表B中包含分级级别的列
    new_col                新表中分级列的列名
    sep                    分隔符
    keep_col               是否保留原df需要分割的列（故障所在列）
    
    样例：
    df_todo

  车辆ID        仓库          省份      城市            出库时间              出库日期       车型       操作人                               配件明细  
7556111144   石岩1号仓库      广东省   深圳市     2017-12-11 14:10:05         2017-12-11    Lite     刘俊峰                车轮组-后轮组(带飞轮/保护盖/外胎)  
7556510815   罗湖清水河仓库   广东省    深圳市    2017-12-11 16:25:41         2017-12-11    Lite2.0  陈伟斌                          车把装置-车首碗组  
8620890320   塘朗2号仓库      广东省   深圳市     2017-12-11 20:25:06         2017-12-11    Lite2.0  叶俊廷      传动装置-脚踏(左)|传动装置-曲柄(不带大齿盘)|传动装置-飞轮    
...    
    
    df_set
    

 车型           配件明细         分级
Classic       传动系统-中轴        3     
Classic       传动系统-传动轴      3   
Classic       传动系统-右曲柄      1         
Classic       传动系统-右脚踏      1      
...     
    
    
    classify_by_A_others=['车型']
    
    
    col_to_classify='配件明细'
    
    
    classify_by_B = ['车型','配件明细']
    
    classify_target ='分级'
    
    new_col = '分级'
    
    sep='|'
    
    keep_col = True
    
    
    
    输出：
    new df
    
  车辆ID       仓库            省份     城市            出库时间               出库日期      车型      操作人                                 配件明细                       分级  
7556111144   石岩1号仓库      广东省   深圳市    2017-12-11 14:10:05          2017-12-11    Lite      刘俊峰                车轮组-后轮组(带飞轮/保护盖/外胎)                 2.0  
7556510815  罗湖清水河仓库    广东省   深圳市    2017-12-11 16:25:41          2017-12-11    Lite2.0    陈伟斌                          车把装置-车首碗组                      2.0  
8620890320   塘朗2号仓库      广东省   深圳市    2017-12-11 20:25:06          2017-12-11    Lite2.0    叶俊廷       传动装置-脚踏(左)|传动装置-曲柄(不带大齿盘)|传动装置-飞轮   2.0  
8621716833   塘朗1号仓库      广东省   深圳市    2017-12-11 16:42:30          2017-12-11    Lite2.0    王伟卫                            坐垫装置-座垫                        1.0      
    
    程序设计思路：
    将需要分割列按照分隔符分开，每个元素生成一个新列，分级，然后取每一行的最大值。
    
    """

    
    #######
#    a = [df_good,failure_set_clean,[Biketype],Failure,[Biketype,Failure],Devision,Devision,'|',True]
#    
#    df_todo = a[0]
#    df_set = a[1]
#    classify_by_A_others =a[2]
#    col_to_classify = a[3]
#    classify_by_B = a[4]
#    classify_target = a[5]
#    new_col = a[6]
#    sep = a[7]
#    keep_col =a[8]
#    
    #######
    
    #修改过的函数，会返回一组列名供之后使用
    #先将故障按分隔符分开，成为新的列
    def seperate_col(df_todo,col,sep="，",trans=False,keep_col=True): 
        data = df_todo.copy()
        
        #转换为list
        data[col] = data[col].str.split(sep)
        num_col = 'num'
        data[num_col] = data[col].apply(len)
        
        for i in range(max(data[num_col])):
            name=num_col + '_'+str(i+1)
            data[name]=data[col].apply(lambda x:x[i] if (len(x)>i) else '')
        
        num_list=[num_col+'_'+str(i+1) for i in range(max(data[num_col]))]
        if keep_col:
            data.drop([num_col],axis=1,inplace=True)
            data[col]= df_todo[col]
        else:
            data.drop([col,num_col],axis=1,inplace=True)
        if trans:
            data = trans_col(data,num_list,col)
        return data,num_list


    ###########
    data,num_list = seperate_col(df_todo,col_to_classify,sep=sep,trans=False,keep_col=keep_col)
    columns = data.columns
    classify_set = df_set.loc[:,classify_by_B+[classify_target]]
    
    #然后每列进行合并、匹配，给出一个分级，并替换原列。  未匹配到的故障get blank value

    for col in num_list:
        ####
#        col='num_1'
        ####  
        
        #给需要合并的列改列名
        temp_classify_by_B_name = [col+'_y' for col in classify_by_B]
        temp_target_name = col+'_y'
        
        old_columns_name = list(classify_set.columns)
        
        classify_set.columns = temp_classify_by_B_name + [temp_target_name]
        
        data =  pd.merge(data,classify_set, left_on=classify_by_A_others+[col], right_on=temp_classify_by_B_name, how='left')
        
        classify_set.columns = old_columns_name
        """
        Notice #1
        """
        data = data.loc[:,list(columns)+[temp_target_name]].fillna("0")
        data[col] = data[temp_target_name].astype(int)
        del data[temp_target_name]
    
    

    
    data[new_col] = data.loc[:,num_list].max(axis=1)
    
    for col in num_list:
        del data[col]

    return data


	
#修改列名，应对嵌套列名
def rename_complex_col(df_todo):
    """
    有些Excel表格具有二级标题，使用pandas读取时只能读取到部分标题。df的第一列中往往有列名信息，
    该程序会检查第一列是否为空值，若不为空则命名为当前列的列名，遍历之后删除第一行
    
    old df：
        日期       你是摩拜哪个区的司机？       姓名                      地理位置                                 Unnamed: 4   Unnamed: 5  
0                                                                         地址                                     经度         纬度  
1    2017-11-06         南山区                刘X贵            广东省深圳市南山区南头街道桃花园(南山大道)         113.923321   22.532531  
2    2017-11-06         南山区                 李X飞             广东省深圳市南山区南山街道德源花园                113.91829    22.518216  
3    2017-11-06         福田区                刘X鹏            广东省深圳市福田区福田街道益田路免税商务大厦       114.056495   22.537452  
4    2017-11-06         龙华区                 崔XX             广东省深圳市龙华区白石龙路汇龙苑                  114.041927   22.602254  


    new df:
        日期       你是摩拜哪个区的司机？       姓名                        地址                                    经度        纬度                                                                                                                      
1    2017-11-06         南山区                刘X贵            广东省深圳市南山区南头街道桃花园(南山大道)         113.923321   22.532531  
2    2017-11-06         南山区                 李X飞             广东省深圳市南山区南山街道德源花园                113.91829    22.518216  
3    2017-11-06         福田区                刘俊鹏            广东省深圳市福田区福田街道益田路免税商务大厦       114.056495   22.537452  
4    2017-11-06         龙华区                 崔XX             广东省深圳市龙华区白石龙路汇龙苑                  114.041927   22.602254  
    
    """
    data = df_todo.copy()
    col = []
    for j in range(data.shape[1]):
        if  data.iloc[0,j] == "":
            col.append(data.columns[j])
        else:
            col.append(data.iloc[0,j])
    data.columns = col
    data.drop(0,inplace=True)
    data.index = range(data.shape[0])
    return data
	
	
#获取所选dataframe所选范围的值的集合	
def getSetFromDf(df_todo):
    """
    注：可能有问题，慎用
    """
    data = df_todo.copy()
    target_set = set()
    data.applymap(lambda x:target_set.add(x) if x !='' else None)
    return target_set
	
	
	
#清洗车辆id
def match_id(x,regex='^[0-9]{10}$'):
    """
    读入string，抽取其中10位数字；如果string不为空，仍没抽取到，则返回空值
    """
    m = re.search(regex,x)
    if m:
        return m.group(0)
    elif x != "":
        print(x)
        return ""
    else:
        return ""
		

#将多个dataframe写到同一个Excel worksheet里，按行		
def multiple_dfs_by_row(df_list, sheets, file_name, spaces,index=False):
    """
    将多个dataframe间隔一定宽度，写到同一个Excel worksheet里，按行		
    注：生成的文件会覆盖原Excel表。
    
    输入：
    df_list       目标dataframe集合
    sheets        需要写入的worksheet名
    file_name     生成的文件名
    spaces        间隔的宽度
    index         写入的df是否保留index，默认为False，但在写入pivot_table的时候需要保留
    """
    
    writer = pd.ExcelWriter(file_name,engine='xlsxwriter')   
    row = 0
    for dataframe in df_list:
        dataframe.to_excel(writer,sheet_name=sheets,startrow=row , startcol=0,index=index)   
        row = row + len(dataframe.index) + spaces + 1

    writer.save()

#将多个dataframe写到同一个Excel worksheet里，按列
def multiple_dfs_by_col(df_list, sheets, file_name, spaces,startrow=0,tablename=[],index=False):
    """
    将多个dataframe间隔一定宽度，写到同一个Excel worksheet里，按行列	
    注：生成的文件会覆盖原Excel表。
    
    输入：
    df_list       目标dataframe集合
    sheets        需要写入的worksheet名
    file_name     生成的文件名
    spaces        间隔的宽度
    startrow      写入起始的行数
    tablename     写入表的表名list，格式为：大写，斜体，加粗
    index         写入的df是否保留index，默认为False，但在写入pivot_table的时候需要保留
    """
    writer = pd.ExcelWriter(file_name,engine='xlsxwriter')   
    
    col = 0
    for i,dataframe in enumerate(df_list):
        dataframe.to_excel(writer,sheet_name=sheets,startrow=startrow , startcol=col,index=index)
        if len(tablename) > 0:
            workbook  = writer.book
            worksheet = writer.sheets[sheets] 
            cell_format = workbook .add_format({'bold': True, 'italic': True,'size':30})
            worksheet.write_string(startrow-1,col,tablename[i],cell_format)
   
        col = col + len(dataframe.columns) + spaces + 1
    writer.save()
	
	


#err_log生成器，使用于日常频繁运行的程序
def log_generator(err_name,err_detail,filename,path='D:\\error_log\\'):
    """
    自动生成错误日志log到path，需要输入的错误内容为err_detail
    
    输入：
    err_name       错误名，自己定义
    err_detail     错误详细信息
    filename       生成的日志文件名
    path           日志文件保存路径
    
    """
    
    today = datetime.date.today()
    today_str = today.strftime('%Y_%m_%d')
    #输出文件名
    output_filename = filename + '_' +today_str+'.txt'
    #输出路径
    outdir = path + output_filename 
    with open(outdir, "w") as output:
        output.write(err_name+":")
        output.write(str(err_detail))
    print(err_detail)

#获取列的factor
def getFactors(df,col):
    """
    获取dataframe某列的唯一值
    """
    return list(set(df[col]))

#取差集A-B
def get_subtract_list(A,B):
    """
    输入两个list或set，返回A去除B中元素后的差集
    """
    return list(set(A).difference(set(B)))

#list取差集
def getDifference(origin_list,item_to_remove=[]):
    """
    输入两个list或set，返回A去除B中元素后的差集
    """
    return list(set(origin_list).difference(set(item_to_remove)))



#按照标识列成为组合列
def trans_mark_value_to_col(df_todo,mark_col,col_to_mark,sep='-',fillna=''):
    """
    按照标识列成为组合列,长数据转化为宽数据，可操作多列
    
    df_todo                需要操作的dataframe
    col_to_keep            需要保留的列的集合
    mark_col               用来转化的列
    col_to_mark=['数量']   需要和mark_col结合的列
    sep                    新列名结合时的连接符
    fillna                 填充None的值
    
    输入：
    old df
    出库日期       分级       数量
    2017-12-11     0.0        2
    2017-12-11     1.0       1931
    2017-12-11     2.0       1452
    2017-12-11     3.0        98
    2017-12-12     0.0        6
    2017-12-12     1.0       2284
    2017-12-12     2.0       1378
    2017-12-12     3.0        82
    
    col_to_keep=['出库日期']
    mark_col='分级'
    col_to_mark=['数量']
    sep='-'
    fillna=''
    
    new df
    出库日期      0.0-数量  1.0-数量  3.0-数量  2.0-数量
    2017-12-11       2       1931       98      1452    
    2017-12-12       6       2284       82      1378    
    """
    
    ################################################
#    df_todo = df_kaoqin_by_chexing.copy()
#    
#    mark_col = 'chexing'
#    
#    col_to_mark = ['车数']
#    
#    sep = '-'
#    
#    fillna=''
    ################################################
    df = df_todo.copy()
        
    #需要先生成一个全组合列
    full_combination_list = list(df.columns)
    
    for mark in list(set(df[mark_col])):
        for col in df[col_to_mark]:
            full_combination_list.append(mark+sep+col)
    
    for col in df[col_to_mark]:
        full_combination_list.remove(col)
    full_combination_list.remove(mark_col)
    
    
    df_new = pd.DataFrame(columns=full_combination_list)
    
    for i in range(df.shape[0]):
    #pop last row of df to dict   
        row_dict = dict(df.iloc[-1,:])  
        df = df[:-1]
        
        
        for col in col_to_mark:
            row_dict[row_dict[mark_col]+sep+col] = row_dict[col]
            del row_dict[col]
        del row_dict[mark_col]
        
        list_dict = {}
        for key,value in row_dict.items():  
            list_dict[key] = [value]
        
        df_temp = pd.DataFrame.from_dict(list_dict, orient='columns')
            
        df_new = df_new.append(df_temp)
    
    df_new.fillna(fillna,inplace=True)
    df_new.index = range(df_new.shape[0])
    
    df_new = df_new.loc[:,full_combination_list]
    return df_new



#按照标识列成为组合列,并纵向合并,会自动删除重复行
def trans_mark_value_to_col_V2(df_todo,col_to_keep,mark_col,col_to_mark,sep='-',fillna='',astype='str'):
    """
    加强版，按照标识列成为组合列，长数据转化为宽数据，自动删除重复行，可操作多列，并指定数值类型
    
    df_todo                需要操作的dataframe
    col_to_keep            需要保留的列的集合
    mark_col               用来转化的列
    col_to_mark=['数量']   需要和mark_col结合的列
    sep                    新列名结合时的连接符
    fillna                 填充None的值
    astype                 是否转换数值，可选'str','float','int'，如果不是str，无效值会被转化为数值0
    
    输入：
    old df
    出库日期       分级       数量
    2017-12-11     0.0        2
    2017-12-11     1.0       1931
    2017-12-11     2.0       1452
    2017-12-11     3.0        98
    2017-12-12     0.0        6
    2017-12-12     1.0       2284
    2017-12-12     2.0       1378
    2017-12-12     3.0        82
    
    col_to_keep=['出库日期']
    mark_col='分级'
    col_to_mark=['数量']
    sep='-'
    fillna=''
    astype='int'
    
    new df
    出库日期      0.0-数量  1.0-数量  3.0-数量  2.0-数量
    2017-12-11       2       1931       98      1452    
    2017-12-12       6       2284       82      1378    
    """
    
    
    ################################################
#    
#    lst =[df,col_to_keep,mark_col,col_to_mark,'-','','float']
#    
#    
#    df_todo = lst[0]
#    
#    col_to_keep = lst[1]
#    
#    mark_col = lst[2]
#
#    
#    col_to_mark = lst[3]
#    
#    sep = lst[4]
#    
#    fillna = lst[5]
#    
#    astype = lst[6]
    ################################################
    df = df_todo.loc[:,col_to_keep+[mark_col]+col_to_mark].copy()
    
    
    #按组合列删除重复列
    df = df.drop_duplicates(col_to_keep+[mark_col], keep='last')

      
    #需要先生成一个全组合列
    full_combination_list = list(df.columns)
    
    mark_list = list(set(df[mark_col]))
    appended_list = []
    
    for mark in mark_list:
        for col in df[col_to_mark]:
            full_combination_list.append(mark+sep+col)
            appended_list.append(mark+sep+col)
    
    
    for col in df[col_to_mark]:
        full_combination_list.remove(col)
    full_combination_list.remove(mark_col)
    
    
    df_new = pd.DataFrame(columns=full_combination_list)
    
    
    #从最后一列开始，一步步的执行
    for i in range(df.shape[0]):
        
    #pop last row of df to dict   
        row_dict = dict(df.iloc[-1,:])  
        df = df[:-1]
        
        
        for col in col_to_mark:
            row_dict[row_dict[mark_col]+sep+col] = row_dict[col]
            del row_dict[col]
        del row_dict[mark_col]
        
        #value转换成list
        list_dict = {}
        for key,value in row_dict.items():  
            list_dict[key] = [value]
        
        df_temp = pd.DataFrame.from_dict(list_dict, orient='columns')
            
        df_new = df_new.append(df_temp)
    
    df_new.fillna(fillna,inplace=True)
    df_new.index = range(df_new.shape[0])
    
    df_new = df_new.loc[:,full_combination_list]
    
    
    #按col_to_keep合并
    df_new['__mark__'] = ""
    for col in col_to_keep:
        df_new['__mark__'] = df_new['__mark__'] + df_new[col]
        
        
    col_to_add = appended_list.copy()
        
    df_wanted = pd.DataFrame(columns=col_to_keep+col_to_add)
    
    for item in list(set(df_new['__mark__'])):
        df_sub = subDataframe(df_new,'__mark__',item)
        df_temp = df_sub.copy()
        
        dict_temp = {}
        
        #需要保留的列
        for col in col_to_keep:
            dict_temp[col] = list(set(df_temp[col].astype(str)))
        
        #需要纵向合并的列
        for col in col_to_add:
            dict_temp[col] = ["".join(df_temp[col].astype(str))]

        df_temp = pd.DataFrame.from_dict(dict_temp, orient='columns')
        
        df_wanted = df_wanted.append(df_temp)
    df_wanted.index = range(df_wanted.shape[0])
    df_wanted = df_wanted.loc[:,col_to_keep+col_to_add]
    
    
    def trans_float(x):
        try:
            x = float(x)
            return x
        except:
            return 0
        
    def trans_int(x):
        try:
            x = int(x)
            return x
        except:
            return 0
        
    #改变输出格式    
    if astype =='float' or astype == 'int':
        for col in col_to_add:

            df_wanted[col] = df_wanted[col].apply(lambda x:trans_float(x))
            if astype =='int':
                df_wanted[col] = df_wanted[col].apply(lambda x:trans_int(x))
        #df_wanted.loc[:,col_to_add] = df_wanted.loc[:,col_to_add].applymap(lambda x:trans_float(x))
    return df_wanted



def trans_float(x):
    """
    将目标值转化为float，如果转换失败，则返回0
    """
    try:
        x = float(x)
        return x
    except:
        return 0
    
def trans_int(x):
    """
    将目标值转化为int，如果转换失败，则返回0
    """
    try:
        x = int(x)
        return x
    except:
        return 0




#将行中的值转化成列,并纵向合并,会自动删除重复行，不会
def trans_mark_value_to_col_V3(df_todo,col_to_keep,mark_col,col_to_mark,fillna='',astype='str',reverse=False):
    """
    超级加强版，按照标识列成为组合列，长数据转化为宽数据，自动删除重复行，可操作多列，并指定数值类型，可指定新列名的连接顺序
    
    df_todo                需要操作的dataframe
    col_to_keep            需要保留的列的集合
    mark_col               用来转化的列
    col_to_mark=['数量']   需要和mark_col结合的列
    sep                    新列名结合时的连接符
    fillna                 填充None的值
    astype                 是否转换数值，可选'str','float','int'
    reverse                新列名连接顺序     '0.0-数量'    '数量-0.0'
    
    输入：
    old df
    出库日期       分级       数量
    2017-12-11     0.0        2
    2017-12-11     1.0       1931
    2017-12-11     2.0       1452
    2017-12-11     3.0        98
    2017-12-12     0.0        6
    2017-12-12     1.0       2284
    2017-12-12     2.0       1378
    2017-12-12     3.0        82
    
    col_to_keep=['出库日期']
    mark_col='分级'
    col_to_mark=['数量']
    sep='-'
    fillna=''
    astype='int'
    reverse=False
    
    new df
    出库日期      0.0-数量  1.0-数量  3.0-数量  2.0-数量
    2017-12-11       2       1931       98      1452    
    2017-12-12       6       2284       82      1378    
    """
    
    ################################################
#    
#    lst =[df,col_to_keep,mark_col,col_to_mark,'-','','float']
#    
#    
#    df_todo = lst[0]
#    
#    col_to_keep = lst[1]
#    
#    mark_col = lst[2]
#
#    
#    col_to_mark = lst[3]
#    
#    sep = lst[4]
#    
#    fillna = lst[5]
#    
#    astype = lst[6]
    ################################################
    sep='_$_'
    
    if type(col_to_mark) != list:
        temp = []
        temp.append(col_to_mark)
        col_to_mark = temp
    
    
    df = df_todo.loc[:,col_to_keep+[mark_col]+col_to_mark].copy()
    
    
    #按组合列删除重复列
    df = df.drop_duplicates(col_to_keep+[mark_col], keep='last')

      
    #需要先生成一个全组合列
    full_combination_list = list(df.columns)
    
    mark_list = list(set(df[mark_col]))
    appended_list = []
    
    for mark in mark_list:
        for col in df[col_to_mark]:
            full_combination_list.append(mark+sep+col)
            appended_list.append(mark+sep+col)
    
    
    for col in df[col_to_mark]:
        full_combination_list.remove(col)
    full_combination_list.remove(mark_col)
    
    
    df_new = pd.DataFrame(columns=full_combination_list)
    
    
    #从最后一列开始，一步步的执行pop
    for i in range(df.shape[0]):
        
    #pop last row of df to dict   
        row_dict = dict(df.iloc[-1,:])  
        df = df[:-1]
        
        
        for col in col_to_mark:
            row_dict[row_dict[mark_col]+sep+col] = row_dict[col]
            del row_dict[col]
        del row_dict[mark_col]
        
        #value转换成list
        list_dict = {}
        for key,value in row_dict.items():  
            list_dict[key] = [value]
        
        df_temp = pd.DataFrame.from_dict(list_dict, orient='columns')
            
        df_new = df_new.append(df_temp)
    
    df_new.fillna(fillna,inplace=True)
    df_new.index = range(df_new.shape[0])
    
    df_new = df_new.loc[:,full_combination_list]
    
    
    #按col_to_keep合并
    df_new['__mark__'] = ""
    for col in col_to_keep:
        df_new['__mark__'] = df_new['__mark__'] + df_new[col]
        
        
    col_to_add = appended_list.copy()
        
    df_wanted = pd.DataFrame(columns=col_to_keep+col_to_add)
    
    for item in list(set(df_new['__mark__'])):
        df_sub = subDataframe(df_new,'__mark__',item)
        df_temp = df_sub.copy()
        
        dict_temp = {}
        
        #需要保留的列
        for col in col_to_keep:
            dict_temp[col] = list(set(df_temp[col].astype(str)))
        
        #需要纵向合并的列
        for col in col_to_add:
            dict_temp[col] = ["".join(df_temp[col].astype(str))]

        df_temp = pd.DataFrame.from_dict(dict_temp, orient='columns')
        
        df_wanted = df_wanted.append(df_temp)
    df_wanted.index = range(df_wanted.shape[0])
    df_wanted = df_wanted.loc[:,col_to_keep+col_to_add]
    
    def trans_float(x):
        try:
            x = float(x)
            return x
        except:
            return 0
        
    def trans_int(x):
        try:
            x = int(x)
            return x
        except:
            return 0
        
    #改变输出格式    
    if astype =='float' or astype == 'int':
        for col in col_to_add:
            df_wanted[col] = df_wanted[col].apply(lambda x:trans_float(x))
            if astype =='int':
                df_wanted[col] = df_wanted[col].apply(lambda x:trans_int(x))
        #df_wanted.loc[:,col_to_add] = df_wanted.loc[:,col_to_add].applymap(lambda x:trans_float(x))
    
    
    #将列去除mark并排序
    df_wanted.columns = [x.replace(sep+col_to_mark[0],"") for x in df_wanted.columns]

    new_col_list = list(set(df_wanted.columns).difference(set(col_to_keep)))
    
    new_col_list = sorted(new_col_list,reverse=reverse)


    df_wanted = df_wanted.loc[:,col_to_keep+new_col_list]
        
    return df_wanted




	
#将python生成的pivot表单汇总	
def sum_pivot(df_todo,name = '总计'):
    """
    将python生成的pivot_table加上列汇总和行汇总
    """
    #################
#   df_todo = df_quyu.copy()
#        
#   name = '总计'
    #################
    
    df = df_todo.copy()
    
    #列汇总
    if len(df.index)>1:
        temp_series = df.sum()
        
        name_list =[name]
        if type(df.index[0]) != str:
            for item in list(df.index[0])[:-1]:
                name_list.append('')
        
        temp_series.name = tuple(name_list)
        df = df.append(temp_series)
    
    
    if len(df.columns)>1:
        #行汇总
        df = df.T
        temp_series =df.sum()
        
        temp_list = list(df.index[0])
        temp_list.pop()
        temp_list.append(name)
        
        temp_series.name = tuple(temp_list)
        df = df.append(temp_series)
        
        df = df.T
    
    return df


#将excel 的数字格式日期转换为字符串或python日期格式
def excel_dateTranslation(num,trans='str'):
    """
    将excel 的数字格式日期转换为字符串或python日期格式
    """
    from datetime import datetime
    
    excel_date = num
    dt = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + excel_date - 2)     
    
    if trans=='str':
        return str(dt)[0:10]
    else:
        return dt
    

#获取Excel的列名和行名，返回str或tuple    
def get_excel_index(row,col,ws,trans='str'):
    """
    按行列号获取excel单元格的index值，返回string如'A1'或tuple如('A','1')
    ws是工作表对象，需要配合openpyxl来使用。
    例：
    wb = load_workbook(os.path.join(DIR,file_fill_in))
    ws = wb.get_sheet_by_name(quyu)
    """
    #########
#    row = 4
#    col = 4
    #########
    cell_name = str(ws.cell(row=row,column=col))
    dot = cell_name.find(".")
    cell_index = cell_name[dot+1:-1]
    
    if trans=='str':
        return cell_index
    elif trans == 'tuple':
        col_name = re.compile(r'^[A-Z]+').search(cell_index).group()
        row_name = re.compile(r'\d+$').search(cell_index).group()
        return (col_name,row_name)
    
#通过位置获取Excel Value   
def get_excel_value(row,col,ws):
    """
    按行列号获取excel单元格的值，ws是工作表对象，需要配合openpyxl来使用。
    例：
    wb = load_workbook(os.path.join(DIR,file_fill_in))
    ws = wb.get_sheet_by_name(quyu)
    """
    return ws.cell(row=row,column=col).value





#df 纵向汇总
def sum_df_col(df_todo,col_not_sum=[],name='总计'):
    """
    将python生成的pivot_table加上列汇总
    
    输入：
    df_todo        需要汇总的dataframe
    col_not_sum    不需要汇总的列名的集合，如日期、人们
    name           汇总栏的index名
    """
    
    df = df_todo.copy()
    
    if type(col_not_sum) != list:
        col_not_sum = [] + [col_not_sum]
        
    col_to_sum = getDifference(df.columns,col_not_sum)
    
    #列汇总
    temp_series = df.loc[:,col_to_sum].sum()
    temp_series.name = name
    
    for col in col_not_sum:
        temp_series[col] = '-'
        
    df = df.append(temp_series)
    
    return df


#df 横向汇总
def sum_df_row(df_todo,col_not_sum,name='总计'):
    """
    将python生成的pivot_table加上行汇总
    
    
    输入：
    df_todo        需要汇总的dataframe
    col_not_sum    不需要汇总的列名的集合，如日期、人们
    name           汇总栏的index名
    """
    
    ######
#    df_todo = a
#    col_not_sum = Date
#    name='总计'
    ######
    
    df = df_todo.copy()
    
    df = df.T
    
    if type(col_not_sum) != list:
        col_not_sum = [] + [col_not_sum]
        
    col_to_sum = getDifference(df.index,col_not_sum)
    
    df_to_sum = df.loc[col_to_sum,:]
    
    df_seped  = df.loc[col_not_sum,:]
    
    df_sumed = df_to_sum.sum()
    
    df_sumed.name = name
    
    df_wanted = df_seped.append(df_to_sum).append(df_sumed)
    
    df_wanted = df_wanted.T
    
    #列排序
    df_wanted = df_wanted.loc[:,list(df_todo.columns)+[name]]
    #转换float格式
    for col in col_to_sum+[name]:
        df_wanted[col] = df_wanted[col].astype(float)
    
    return df_wanted
 
    
#df双向汇总
def sum_df(df_todo,col_not_sum,name='总计'):
    """
    双向汇总dataframe
    
    输入：
    df_todo        需要汇总的dataframe
    col_not_sum    不需要汇总的列名的集合，如日期、人们
    name           汇总栏的index名
    
    
    """
    
    df = df_todo.copy()
    temp = sum_df_col(df,col_not_sum,name)
    temp = sum_df_row(temp,col_not_sum,name)
    return temp

def return_bool(x,regex='^[0-9]{10}$'):
    """
    输入string，按正则匹配，返回bool值
    """
    m = re.search(regex,x)
    if m:
        return True
    else:
        print(x)
        return False
    
    
    
    
    
#找出df1中有，df2中没有的组合列
def get_different_col_combination(df_1,df_2,col_to_compare1,col_to_compare2=[]):
    """
    找出df1中有，df2中没有的组合列
    
    输入：
    df_1                            dataframe A
    df_2                            dataframe B
    col_to_compare1                 A中用来判断的列的组合
    col_to_compare2                 B中用来判断的列的组合，如果不指定，则默认和A一致
    
    df_1
    国家                姓名                武力值
    蜀国               诸葛亮                 0
    蜀国                关羽                   99
    魏国                曹操                   56
    
    
    df_2
    
    国家                姓名                智力值
    蜀国               诸葛亮                 120
    魏国                曹操                   100
    
    
    col_to_compare1  = ['国家','姓名']
    col_to_compare2  = ['国家','姓名']
    
    new df:
    国家                姓名 
    蜀国                关羽 
    """

#==============================================================================
#lst = [df_sep,failure_set,[Failure,Biketype]] 
#df_1 =lst[0]
#df_2 =lst[1]
#col_to_compare1 =lst[2]
#col_to_compare2=[]
#==============================================================================
    
    if type(col_to_compare1) != list:
        col_to_compare1 = [].append(col_to_compare1)
    
    if col_to_compare2==[]:
        col_to_compare2 = col_to_compare1
    
    df1 = df_1.loc[:,col_to_compare1].copy()
    df2 = df_2.loc[:,col_to_compare2].copy()
    
    df1['__mark__'] = ''
    
    for col in col_to_compare1:
        df1['__mark__'] = df1['__mark__']+df1[col]
        
    df2['__mark__'] = ''
    
    for col in col_to_compare2:
        df2['__mark__'] = df2['__mark__']+df2[col]
        
    set_mark_df2 = set(df2['__mark__'])
    
    df1['__mark__'] = df1['__mark__'].apply(lambda x:1 if x in set_mark_df2 else 0)
    
    df_out = subDataframe(df1.loc[:,col_to_compare1+['__mark__']],'__mark__',0)
    
    del df_out['__mark__']
    
    df_out = df_out.drop_duplicates(col_to_compare1, keep='last')
    
    return df_out
    

#str(datetime.datetime.now())[0:-7].replace(" ","_").replace(":","")


#读取脏格式的csv文件
def readcsvV1(filename,returnExceptions=False):
    """
    本函数适用于雅典娜下载的车辆原始csv文件，如果csv文件下载完整，且包含最后一行的相关状态信息，
    程序会以utf8的格式读取，并自动去除列名中的特殊符号如'\t','?'
    如果以utf8格式读取失败，程序会自动转到ANSI格式读取，并将错误行生成为df_exception;如果returnExceptions=True，程序会返回df和df_exception
    """
    import csv
    try:
        print('\nutf8格式读取...')
        with open(filename, newline='',encoding="utf8") as csvfile:
           
            data_list = []
            
            spamreader = csv.DictReader(csvfile, delimiter=',', quotechar='|')  
            
            for i,row in enumerate(spamreader):
                
                #print(row)
                data_list.append(row)
                
        df = pd.DataFrame(data_list)
        
        #删除空行
        col_list = df.columns.copy()
        for i,col in enumerate(col_list):
            if type(col) == type(None):
                del df[col]
        #删除值中的tab
        for col in df.columns:
            df[col] = df[col].apply(lambda x:x.replace("\t",""))
            
        #清洗列名
        for item in [' ','\t','?',',']:
            df.columns = [x.replace(item,'') for x in df.columns]
        
        return df
    
                
    except UnicodeDecodeError as err:
        print('\nutf-8格式读取失败，尝试ANSI...')
                
        with open(filename, newline='',encoding="ANSI") as csvfile:
        
            #spamreader = csv.DictReader(csvfile, delimiter=',', quotechar='|')
            spamreader = csv.reader(csvfile, delimiter=',', quotechar='|')
    
    
    
            #依然使用字典
            #如果和标题匹配，则确认；如果不匹配，另外生成dataframe做处理
            data_list = []
            exception_list = []
            for i,row in enumerate(spamreader):
    
                #print(row)
                #强行分割
                "###具体的分隔方式可能需要根据实际情况修改###"
                quotation_removed = row[0].replace('"','\t')
                row_splited = quotation_removed.split('\t')   
                row_splited = list(filter(None, row_splited))    
                
                #截取并清洗列名
                if i == 0:
                    columns = row_splited
                    for item in [' ','\t','?',',']:
                        columns = [x.replace(item,'') for x in columns]
    
                #将正常数据和例外分开
                elif len(row_splited)==len(columns):
                    temp_dict = dict()
                    for i in range(len(columns)):
                        temp_dict[columns[i]]=row_splited[i]               
                    data_list.append(temp_dict)
                else:
                    exception_list.append(row_splited)
    
            df=pd.DataFrame(data_list)
            df = df.loc[:,columns]
            
            df_exception = pd.DataFrame(exception_list)
            
            
            print("df length        : ",df.shape[0])
            print("exception length : ",df_exception.shape[0])
            
            #返回好的值和例外
            if returnExceptions:
                return df,df_exception
            else:
                return df



##封装excel类
#from openpyxl import load_workbook  
#from openpyxl import Workbook  
#from openpyxl.chart import BarChart, Series, Reference, BarChart3D  
#from openpyxl.styles import Color, Font, Alignment  
#from openpyxl.styles.colors import BLUE, RED, GREEN, YELLOW  
#  
#class Write_excel(object):  
#    def __init__(self,filename):  
#        self.filename = filename  
#        self.wb = load_workbook(self.filename)  
#        self.ws = self.wb.active  
#  
#    def write(self, coord, value):  
#        # eg: coord:A1  
#        self.ws.cell(coord).value = value  
#        self.wb.save(self.filename)  
#  
#    def merge(self, rangstring):  
#        # eg: rangstring:A1:E1  
#        self.ws.merge_cells(rangstring)  
#        self.wb.save(self.filename)  
#  
#    def cellstyle(self, coord, font, align):  
#        cell = self.ws.cell(coord)  
#        cell.font = font  
#        cell.alignment = align  
#  
#    def makechart(self, title, pos, width, height, col1, row1, col2, row2, col3, row3, row4):  
#        ''''':param title:图表名 
#                  pos:图表位置 
#                  width:图表宽度 
#                  height:图表高度 
#        '''  
#        data = Reference(self.ws, min_col=col1, min_row=row1, max_col=col2, max_row=row2)  
#        cat = Reference(self.ws, min_col=col3, min_row=row3, max_row=row4)  
#        chart = BarChart3D()  
#        chart.title = title  
#        chart.width = width  
#        chart.height = height  
#        chart.add_data(data=data, titles_from_data=True)  
#        chart.set_categories(cat)  
#        self.ws.add_chart(chart, pos)  
#        self.wb.save(self.filename)  



##逆透视，mark合并成一列	
#def trans_pivot(df_todo,other_col_to_keep,wanted_col,mark_col=[],new_mark_col_name ='mark',new_value_col_name='value' ,sep='-',inverse=True):
#    """
#
#    """
#    ##############
##    df_todo = df_ganyu_sum.copy()
##    other_col_to_keep=['区域']
##    mark_col='工作身份'
##    wanted_col=['故障干预总数','干预人数']
##    new_mark_col_name ='mark'
##    new_value_col_name='value' 
##    sep = '-'
##    inverse=True
#    ##############
#    data = df_todo.copy()
#    df_all = pd.DataFrame(columns=other_col_to_keep+[new_mark_col_name,new_value_col_name])
#    for col in wanted_col:
#        #####
#        #col='故障干预总数'
#        #col='干预人数'
#        #####
#        if len(mark_col)>0: 
#            df_temp = data.ix[:,other_col_to_keep+[mark_col,col]]
#            df_temp.rename(columns={col:new_value_col_name},inplace=True)
#            df_temp[new_mark_col_name]=df_temp[mark_col]
#            del df_temp[mark_col]
#            if inverse:
#                df_temp[new_mark_col_name] = df_temp[new_mark_col_name].apply(lambda x:col+sep+x)
#            else:
#                df_temp[new_mark_col_name] = df_temp[new_mark_col_name].apply(lambda x:x+sep+col)
#            df_all = df_all.append(df_temp)
#        else:
#            df_temp = data.ix[:,other_col_to_keep+[col]]
#            df_temp.rename(columns={col:new_value_col_name},inplace=True)
#            df_temp[new_mark_col_name]= col
#            df_all = df_all.append(df_temp)
#            
#    return df_all.loc[:,other_col_to_keep+[new_mark_col_name,new_value_col_name]]



